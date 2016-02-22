from xlsx import *
from combination import *
from openpyxl import Workbook

#excel based
xldata = []
for col in first_columns:
    xldata.append(xlsxData(start_x=col))
u = U(xldata)

combo_of_vendors = combinations(u.vendors, nbr_of_selected)

#testing
first_row_results = 2
first_col_results = 1
y = 0 #index of combo item
xlsxSheetTitle = str(nbr_of_selected) + '_' + str(nbr_of_vendors)
ws_results = wb.create_sheet(title=xlsxSheetTitle)
ws_results.cell(row=1, column=1).value=1

for combo in combo_of_vendors:
    print('========= run ' + str(y+1) + ' =========')
    combo_name = ''
    for c in combo:
        if combo_name == '':
            combo_name += c
        else:
            combo_name += ' & ' + c
    ws_results.cell(row = first_row_results + (nbr_of_lines + rowbreak) * y, column = first_col_results).value = combo_name

    ui = u.getInterimAtoms(filterlist=[combo])
    ur = u.removeRedundancy(ui)
    us = u.structureAtoms(ur)
    #print(us)
    start = [a for a in us if a[3][0]==1] #starting atoms
    u.getAllCombinations(us, start) #go through iteration

    key = (combo,) #filterlist as key to combination dict
    x = 0 #index of combination in each combo
    for combination in u.getCombinationDict(key):
        print('========= start printing combination ' + str(x + 1) + ' =========')
        for atom in combination:
            for id in atom[3]:
                ws_results.cell(row = first_row_results + id + (nbr_of_lines + rowbreak) * y, column = first_col + (2 + colbreak) * x + 1).value = atom[0]
                ws_results.cell(row = first_row_results + id + (nbr_of_lines + rowbreak) * y, column = first_col + (2 + colbreak) * x + 2).value = atom[1]
            ws_results.cell(row = first_row_results + atom[3][0] + (nbr_of_lines + rowbreak) * y, column = first_col + (2 + colbreak) * x + 3).value = atom[2]
        print('=========   end printing combination ' + str(x + 1) + ' =========')
        x += 1
    y += 1
print("Powered by yuanjingwei.com")
wb.save('Analysis ' + xlsxSheetTitle + '.xlsx')
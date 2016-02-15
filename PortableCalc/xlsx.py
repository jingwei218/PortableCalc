from openpyxl import load_workbook

#open file and get sheet
wb = load_workbook("zzzTest.xlsx")
ws = wb.get_sheet_by_name("Analysis")

#get configuration
nbr_of_mixin = ws.cell(row=3, column=2).value
first_row = ws.cell(row=4, column=2).value
first_col = ws.cell(row=5, column=2).value
rowbreak = ws.cell(row=6, column=2).value
colbreak = ws.cell(row=7, column=2).value
nbr_of_vendors = ws.cell(row=3, column=5).value
nbr_of_incumbents = ws.cell(row=4, column=5).value
nbr_of_lines = ws.cell(row=5, column=5).value
first_row_analysis = ws.cell(row=6, column=5).value
nbr_of_selected = ws.cell(row=7, column=5).value
first_columns = [first_col + 2 * (nbr_of_incumbents + nbr_of_vendors + colbreak) * n for n in range(nbr_of_mixin)]
last_row = ws.max_row
last_col = ws.max_column

#class of data
class xlsxData:
    def __init__(self, start_x=first_columns[0], start_y=first_row, skip_mixin_columns=nbr_of_incumbents, title=True):
        self.skip_mixin_columns=skip_mixin_columns #skip incumbent columns, only include new vendors
        self.start_x = start_x
        self.start_y = start_y
        self.includeTitle = title
        self.table = [None] * nbr_of_vendors
        for i in range(nbr_of_vendors):
            self.table[i] = [None] * nbr_of_lines
            for j in range(nbr_of_lines):
                self.table[i][j] = [None] * 4
                tempCol = self.start_x + self.skip_mixin_columns + i + 1
                tempRow = self.start_y + j + 1
                self.table[i][j][0] = ws.cell(row=self.start_y, column=tempCol).value #names of vendors
                self.table[i][j][1] = ws.cell(row=tempRow, column=tempCol).value if ws.cell(row=tempRow, column=tempCol).value != None else "U" #names of aggregation groups
                self.table[i][j][2] = ws.cell(row=tempRow, column=tempCol+nbr_of_incumbents+nbr_of_vendors+colbreak).value #values of aggregation groups
                self.table[i][j][3] = ws.cell(row=tempRow, column=start_x).value #ids